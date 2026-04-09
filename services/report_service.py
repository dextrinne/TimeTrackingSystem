"""
Сервис формирования отчётов.
ЛР1-Ф4: Генерация сводного отчёта за период.
ЛР1-Ф5: Выгрузка отчётов в XLSX.
"""

from datetime import datetime

from database.db_manager import DatabaseManager


class ReportService:
    """Сервис для генерации и экспорта отчётов."""

    def __init__(self, db_manager: DatabaseManager):
        self.db = db_manager

    def get_timesheet_report(self, timesheet_id):
        """Сформировать сводный отчёт по табелю.
        
        Returns:
            list of dict: данные для отчёта по каждому сотруднику
        """
        rows = self.db.execute_query(
            """SELECT 
                   e.id_employee,
                   e.fio,
                   e.position,
                   e.norm_hours,
                   COALESCE(SUM(CASE WHEN te.type = 'Рабочий день' THEN 1 ELSE 0 END), 0) as workdays,
                   COALESCE(SUM(CASE WHEN te.type = 'Отпуск' THEN 1 ELSE 0 END), 0) as vacations,
                   COALESCE(SUM(CASE WHEN te.type = 'Больничный' THEN 1 ELSE 0 END), 0) as sick_leaves,
                   COALESCE(SUM(CASE WHEN te.type = 'Командировка' THEN 1 ELSE 0 END), 0) as business_trips,
                   COALESCE(SUM(CASE WHEN te.type = 'Неявка' THEN 1 ELSE 0 END), 0) as absences,
                   COALESCE(SUM(te.hours_worked), 0) as total_hours
               FROM employee e
               LEFT JOIN timesheet_entry te ON e.id_employee = te.employee_id AND te.timesheet_id = %s
               GROUP BY e.id_employee, e.fio, e.position, e.norm_hours
               ORDER BY e.fio""",
            (timesheet_id,),
            fetch=True
        )

        report_data = []
        if rows:
            for row in rows:
                report_data.append({
                    'employee_id': row[0],
                    'fio': row[1],
                    'position': row[2],
                    'norm_hours': row[3],
                    'workdays': row[4],
                    'vacations': row[5],
                    'sick_leaves': row[6],
                    'business_trips': row[7],
                    'absences': row[8],
                    'total_hours': float(row[9])
                })

        return report_data

    def get_employee_report(self, employee_id, period_start, period_end):
        """Сформировать отчёт по конкретному сотруднику за период."""
        row = self.db.execute_query(
            """SELECT e.fio, e.position, e.norm_hours
               FROM employee e
               WHERE e.id_employee = %s""",
            (employee_id,),
            fetchone=True
        )

        if not row:
            return None

        entries = self.db.execute_query(
            """SELECT te.date, te.hours_worked, te.type
               FROM timesheet_entry te
               JOIN timesheet t ON te.timesheet_id = t.id_timesheet
               WHERE te.employee_id = %s 
                 AND t.period_start >= %s 
                 AND t.period_end <= %s
               ORDER BY te.date""",
            (employee_id, period_start, period_end),
            fetch=True
        )

        return {
            'fio': row[0],
            'position': row[1],
            'norm_hours': row[2],
            'period_start': period_start,
            'period_end': period_end,
            'entries': entries
        }

    def export_to_xlsx(self, timesheet_id, file_path):
        """Экспорт табеля в XLSX (ЛР1-Ф5).
        
        Args:
            timesheet_id: ID табеля
            file_path: Путь для сохранения файла
        
        Returns:
            bool: True если экспорт успешен
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            # Получаем данные
            report_data = self.get_timesheet_report(timesheet_id)
            
            # Получаем период табеля
            timesheet_row = self.db.execute_query(
                """SELECT period_start, period_end, status 
                   FROM timesheet WHERE id_timesheet = %s""",
                (timesheet_id,),
                fetchone=True
            )

            if not timesheet_row:
                return False

            period_start, period_end, status = timesheet_row

            # Создаём книгу
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Табель рабочего времени"

            # Стили
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, color="FFFFFF", size=10)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Заголовок
            ws.merge_cells('A1:J1')
            title_cell = ws['A1']
            title_cell.value = f"ТАБЕЛЬ УЧЁТА РАБОЧЕГО ВРЕМЕНИ"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:J2')
            period_cell = ws['A2']
            period_cell.value = f"Период: {period_start} — {period_end}"
            period_cell.font = Font(size=11)
            period_cell.alignment = Alignment(horizontal='center')

            ws.merge_cells('A3:J3')
            status_cell = ws['A3']
            status_cell.value = f"Статус: {status}"
            status_cell.font = Font(size=11)
            status_cell.alignment = Alignment(horizontal='center')

            # Заголовки колонок
            headers = [
                '№', 'ФИО', 'Должность', 'Норма часов',
                'Рабочие дни', 'Отпуск', 'Больничный', 
                'Командировка', 'Неявка', 'Всего часов'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Данные
            for row_idx, data in enumerate(report_data, 6):
                row_data = [
                    row_idx - 5,
                    data['fio'],
                    data['position'],
                    data['norm_hours'],
                    data['workdays'],
                    data['vacations'],
                    data['sick_leaves'],
                    data['business_trips'],
                    data['absences'],
                    data['total_hours']
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal='center')

            # Итоговая строка
            if report_data:
                total_row = len(report_data) + 6
                ws.merge_cells(start_row=total_row, start_column=1,
                             end_row=total_row, end_column=3)
                ws.cell(row=total_row, column=1, value="ИТОГО:").font = Font(bold=True)

                ws.cell(row=total_row, column=5,
                       value=sum(d['workdays'] for d in report_data)).font = Font(bold=True)
                ws.cell(row=total_row, column=6,
                       value=sum(d['vacations'] for d in report_data)).font = Font(bold=True)
                ws.cell(row=total_row, column=7,
                       value=sum(d['sick_leaves'] for d in report_data)).font = Font(bold=True)
                ws.cell(row=total_row, column=8,
                       value=sum(d['business_trips'] for d in report_data)).font = Font(bold=True)
                ws.cell(row=total_row, column=9,
                       value=sum(d['absences'] for d in report_data)).font = Font(bold=True)
                ws.cell(row=total_row, column=10,
                       value=sum(d['total_hours'] for d in report_data)).font = Font(bold=True)
            else:
                total_row = 5

            # Ширина колонок
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 12
            for col in ['E', 'F', 'G', 'H', 'I', 'J']:
                ws.column_dimensions[col].width = 12

            # Дата формирования
            ws.merge_cells('A{}:J{}'.format(total_row + 2, total_row + 2))
            ws.cell(row=total_row + 2, column=1,
                   value=f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(italic=True, size=9)

            wb.save(file_path)
            print(f"[OK] Отчёт экспортирован: {file_path}")
            return True

        except Exception as e:
            print(f"[ERROR] Ошибка экспорта в XLSX: {e}")
            return False

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
from pathlib import Path
from typing import Optional

from frontend.styles import DayTypes
from frontend.config import AppConfig


class TemplateExporter:
    """Класс для экспорта табеля в Excel"""
    
    def __init__(self):
        self.last_export_dir = None  # Запоминаем последнюю папку экспорта
    
    def export_timetable(
        self, 
        year: int, 
        month: int, 
        employees: list, 
        day_data: dict, 
        with_stats: bool = True, 
        with_colors: bool = True, 
        save_path: Optional[str] = None
    ) -> Optional[str]:
        """
        Экспорт табеля в Excel
        
        Args:
            year: год
            month: месяц
            employees: список сотрудников
            day_data: данные дней
            with_stats: включать статистику
            with_colors: использовать цвета
            save_path: путь для сохранения (если None - будет предложено выбрать)
        
        Returns:
            str: путь к сохраненному файлу или None если отменено
        """
        # Если путь не указан, запрашиваем у пользователя
        if save_path is None:
            from PyQt6.QtWidgets import QFileDialog
            
            # Предлагаем выбрать папку
            default_name = f"Табель_{AppConfig.MONTHS[month-1]}_{year}.xlsx"
            
            # Используем последнюю папку экспорта или Документы
            initial_dir = self.last_export_dir or str(Path.home() / "Documents")
            
            save_path, _ = QFileDialog.getSaveFileName(
                None,
                "Сохранить табель как",
                str(Path(initial_dir) / default_name),
                "Excel files (*.xlsx);;All files (*.*)"
            )
            
            if not save_path:
                return None  # Пользователь отменил
            
            # Добавляем расширение если не указано
            if not save_path.endswith('.xlsx'):
                save_path += '.xlsx'
            
            # Запоминаем папку для следующего раза
            self.last_export_dir = str(Path(save_path).parent)
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Табель {AppConfig.MONTHS[month-1]} {year}"
        
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Стили
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws.merge_cells(f'A1:{get_column_letter(days_in_month + 2)}1')
        title_cell = ws['A1']
        title_cell.value = f"ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ ЗА {AppConfig.MONTHS[month-1].upper()} {year} г."
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_alignment
        
        # Шапка таблицы
        headers = ['№', 'ФИО', 'Должность']
        for day in range(1, days_in_month + 1):
            headers.append(str(day))
        
        if with_stats:
            headers.extend(['Рабочие', 'Выходные', 'Отпуск', 'Больничный', 'Прочее'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Данные сотрудников
        for emp_idx, employee in enumerate(employees, 1):
            row = emp_idx + 2
            
            # Номер
            cell_num = ws.cell(row=row, column=1, value=emp_idx)
            cell_num.border = border
            cell_num.alignment = center_alignment
            
            # ФИО
            cell_name = ws.cell(row=row, column=2, value=employee.get('name', ''))
            cell_name.border = border
            
            # Должность
            cell_pos = ws.cell(row=row, column=3, value=employee.get('position', ''))
            cell_pos.border = border
            
            # Статистика
            stats = {
                'working_days': 0,
                'weekends': 0,
                'vacation': 0,
                'sick_leave': 0,
                'other': 0,
                'total_hours': 0
            }
            
            # Дни месяца
            for day in range(1, days_in_month + 1):
                day_key = (emp_idx - 1, day)
                cell_data = day_data.get(day_key, {'code': DayTypes.EMPTY, 'hours': 0})
                
                # Поддержка старого формата
                if isinstance(cell_data, str):
                    code = cell_data
                    hours = 0
                else:
                    code = cell_data.get('code', DayTypes.EMPTY)
                    hours = cell_data.get('hours', 0)
                
                name, color, short_code = DayTypes.TYPES[code]
                
                # Для рабочих дней показываем часы
                if code == DayTypes.WORKDAY and hours > 0:
                    display_value = str(hours)
                else:
                    display_value = short_code if code != DayTypes.EMPTY else ""
                
                cell = ws.cell(row=row, column=day + 3, value=display_value)
                cell.alignment = center_alignment
                cell.border = border
                
                if with_colors and code != DayTypes.EMPTY:
                    cell.fill = PatternFill(start_color=color[1:], end_color=color[1:], fill_type="solid")
                
                # Статистика
                if code == DayTypes.WORKDAY:
                    stats['working_days'] += 1
                    stats['total_hours'] += hours
                elif code == DayTypes.WEEKEND:
                    stats['weekends'] += 1
                elif code in [DayTypes.VACATION_MAIN, DayTypes.VACATION_EXTRA]:
                    stats['vacation'] += 1
                elif code == DayTypes.SICK:
                    stats['sick_leave'] += 1
                elif code != DayTypes.EMPTY:
                    stats['other'] += 1
            
            # Добавление статистики
            if with_stats:
                col_offset = days_in_month + 4
                
                cell_wd = ws.cell(row=row, column=col_offset, value=stats['working_days'])
                cell_wd.border = border
                cell_wd.alignment = center_alignment
                
                cell_we = ws.cell(row=row, column=col_offset + 1, value=stats['weekends'])
                cell_we.border = border
                cell_we.alignment = center_alignment
                
                cell_vac = ws.cell(row=row, column=col_offset + 2, value=stats['vacation'])
                cell_vac.border = border
                cell_vac.alignment = center_alignment
                
                cell_sick = ws.cell(row=row, column=col_offset + 3, value=stats['sick_leave'])
                cell_sick.border = border
                cell_sick.alignment = center_alignment
                
                cell_other = ws.cell(row=row, column=col_offset + 4, value=stats['other'])
                cell_other.border = border
                cell_other.alignment = center_alignment
        
        # Итоговая строка
        if with_stats and employees:
            total_row = len(employees) + 3
            cell_total_label = ws.cell(row=total_row, column=1, value="ИТОГО")
            cell_total_label.font = Font(bold=True)
            cell_total_label.border = border
            
            col_offset = days_in_month + 4
            for i, stat in enumerate(['working_days', 'weekends', 'vacation', 'sick_leave', 'other']):
                # Суммируем значения
                total = 0
                for emp_idx in range(len(employees)):
                    row = emp_idx + 3
                    cell_value = ws.cell(row=row, column=col_offset + i).value
                    total += cell_value if cell_value else 0
                
                cell_total = ws.cell(row=total_row, column=col_offset + i, value=total)
                cell_total.font = Font(bold=True)
                cell_total.border = border
                cell_total.alignment = center_alignment
        
        # Автоподбор ширины столбцов
        for col in range(1, len(headers) + 1):
            if col in [2, 3]:  # ФИО и Должность
                ws.column_dimensions[get_column_letter(col)].width = 30
            else:
                ws.column_dimensions[get_column_letter(col)].width = 8
        
        # Дополнительное форматирование
        ws.freeze_panes = 'D3'  # Закрепляем первые три столбца и две строки
        
        # Сохранение файла
        try:
            wb.save(save_path)
            return save_path
        except PermissionError:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(
                None,
                "Ошибка сохранения",
                "Не удалось сохранить файл. Возможно, он открыт в другой программе."
            )
            return None
        except Exception as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(
                None,
                "Ошибка сохранения",
                f"Не удалось сохранить файл:\n{str(e)}"
            )
            return None